# =============================================================================
# src/document_loaders/excel_loader.py
# =============================================================================
"""
Loader para archivos Excel (.xlsx, .xls)
"""
from pathlib import Path
from typing import List, Dict, Any

import openpyxl

from .base_loader import BaseDocumentLoader, DocumentSection, ProcessedDocument


class ExcelLoader(BaseDocumentLoader):
    """Carga y procesa archivos Excel"""

    def __init__(self, max_rows_per_section: int = 100):
        super().__init__()
        self.supported_extensions = {'.xlsx', '.xlsm'}
        self.max_rows_per_section = max_rows_per_section
        # .xls (Excel antiguo) requeriría xlrd, por ahora solo xlsx

    def load(self, file_path: Path, original_filename: str = None) -> ProcessedDocument:
        """Carga un archivo Excel"""
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        # Extract metadata
        metadata = self._extract_metadata(workbook, file_path)

        # Extract sections (una por sheet)
        sections = self.extract_sections(workbook)

        # Generate full content
        full_content = self._generate_full_content(sections)

        # Convertir a ruta relativa
        abs_path = file_path if file_path.is_absolute() else file_path.resolve()
        try:
            relative_path = abs_path.relative_to(Path.cwd())
        except ValueError:
            relative_path = abs_path

        return ProcessedDocument(
            file_path=str(relative_path),
            file_name=file_path.name,
                original_filename=original_filename or file_path.name,
            content=full_content,
            sections=sections,
            metadata=metadata
        )

    def _extract_metadata(self, workbook: openpyxl.Workbook,
                          file_path: Path) -> Dict[str, Any]:
        """Extract metadata from Excel workbook"""
        props = workbook.properties

        metadata = {
            'creator': props.creator or '',
            'title': props.title or '',
            'subject': props.subject or '',
            'description': props.description or '',
            'keywords': props.keywords or '',
            'created': props.created.isoformat() if props.created else '',
            'modified': props.modified.isoformat() if props.modified else '',
            'last_modified_by': props.lastModifiedBy or '',
            'sheet_count': len(workbook.sheetnames),
            'sheet_names': workbook.sheetnames,
            'active_sheet': workbook.active.title if workbook.active else ''
        }

        return metadata

    def extract_sections(self, workbook: openpyxl.Workbook) -> List[DocumentSection]:
        """
        Extract sections from Excel workbook
        Each sheet becomes a section (or multiple if large)
        """
        sections = []

        for sheet in workbook.worksheets:
            # Get sheet data
            data = self._extract_sheet_data(sheet)

            if not data:
                # Empty sheet
                sections.append(DocumentSection(
                    title=f"Sheet: {sheet.title}",
                    content="[Empty sheet]",
                    level=1,
                    metadata={
                        'sheet_name': sheet.title,
                        'row_count': 0,
                        'col_count': 0
                    }
                ))
                continue

            # Detectar si hay headers
            headers = self._detect_headers(data)

            # Si la hoja es pequeña, una sola sección
            if len(data) <= self.max_rows_per_section:
                content = self._format_sheet_data(data, headers)
                sections.append(DocumentSection(
                    title=f"Sheet: {sheet.title}",
                    content=content,
                    level=1,
                    metadata={
                        'sheet_name': sheet.title,
                        'row_count': len(data),
                        'col_count': len(data[0]) if data else 0,
                        'has_headers': bool(headers),
                        'headers': headers
                    }
                ))
            else:
                # Dividir en chunks
                # Primer chunk incluye headers
                for i in range(0, len(data), self.max_rows_per_section):
                    chunk = data[i:i + self.max_rows_per_section]
                    row_start = i + 1
                    row_end = i + len(chunk)

                    # Incluir headers en cada chunk
                    if headers and i > 0:
                        chunk_with_headers = [headers] + chunk
                    else:
                        chunk_with_headers = chunk

                    content = self._format_sheet_data(chunk_with_headers,
                                                      headers if i == 0 else None)
                    sections.append(DocumentSection(
                        title=f"Sheet: {sheet.title} (rows {row_start}-{row_end})",
                        content=content,
                        level=2,
                        metadata={
                            'sheet_name': sheet.title,
                            'row_start': row_start,
                            'row_end': row_end,
                            'row_count': len(chunk),
                            'has_headers': bool(headers),
                            'headers': headers
                        }
                    ))

        return sections

    def _extract_sheet_data(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> List[List]:
        """Extract all data from a sheet"""
        data = []

        for row in sheet.iter_rows(values_only=True):
            # Skip completely empty rows
            if any(cell is not None and str(cell).strip() for cell in row):
                # Convert None to empty string
                clean_row = [str(cell).strip() if cell is not None else '' for cell in row]
                data.append(clean_row)

        return data

    def _detect_headers(self, data: List[List]) -> List[str]:
        """
        Try to detect if first row is headers
        Heuristic: first row has text, rest has more numbers/dates
        """
        if not data or len(data) < 2:
            return []

        first_row = data[0]

        # Check if first row looks like headers
        # (all cells have values and are not too long)
        if all(cell and len(str(cell)) < 100 for cell in first_row):
            # Check if subsequent rows have different types
            has_numbers = False
            for row in data[1:5]:  # Check first few data rows
                for cell in row:
                    if cell and cell.replace('.', '').replace('-', '').isdigit():
                        has_numbers = True
                        break
                if has_numbers:
                    break

            if has_numbers or len(data) > 10:
                return first_row

        return []

    def _format_sheet_data(self, data: List[List],
                           headers: List[str] = None) -> str:
        """Format sheet data as readable text"""
        if not data:
            return "[No data]"

        lines = []

        # Determine column widths for alignment
        col_widths = []
        all_rows = [headers] + data if headers else data

        for col_idx in range(len(all_rows[0]) if all_rows else 0):
            max_width = max(
                len(str(row[col_idx])) if col_idx < len(row) else 0
                for row in all_rows
            )
            col_widths.append(min(max_width, 50))  # Max 50 chars per column

        # Format rows
        start_idx = 1 if headers else 0
        for row in data[start_idx:]:
            formatted_cells = []
            for idx, cell in enumerate(row):
                if idx < len(col_widths):
                    cell_str = str(cell)[:50]  # Truncate long cells
                    formatted_cells.append(cell_str)

            lines.append(' | '.join(formatted_cells))

        # Add header if exists
        if headers:
            header_line = ' | '.join(str(h)[:50] for h in headers)
            separator = '-' * len(header_line)
            lines = [header_line, separator] + lines

        return '\n'.join(lines)

    def _generate_full_content(self, sections: List[DocumentSection]) -> str:
        """Generate full content from sections"""
        full_content = []

        full_content.append("# Excel Workbook")
        full_content.append("")

        current_sheet = None
        for section in sections:
            sheet_name = section.metadata.get('sheet_name', 'Unknown')

            # Add sheet header if new sheet
            if sheet_name != current_sheet:
                full_content.append(f"## {sheet_name}")
                current_sheet = sheet_name

            # Add section content
            if section.level == 2:
                full_content.append(f"### {section.title}")

            full_content.append(section.content)
            full_content.append("")

        return '\n'.join(full_content).strip()
